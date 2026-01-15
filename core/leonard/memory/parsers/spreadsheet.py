"""Parser for spreadsheet files (Excel and CSV)."""

import csv
from pathlib import Path

from leonard.memory.parsers.base import DocumentParser, ParsedDocument, DocumentSection
from leonard.utils.logging import logger


class SpreadsheetParser(DocumentParser):
    """
    Parser for spreadsheet files.

    Extracts data as formatted text.
    Supports: .xlsx, .xls, .csv
    """

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".csv"]

    async def parse(self, file_path: Path) -> ParsedDocument:
        """
        Parse a spreadsheet file.

        Args:
            file_path: Path to the spreadsheet

        Returns:
            ParsedDocument with extracted data
        """
        self._validate_file(file_path)

        ext = file_path.suffix.lower()

        if ext == ".csv":
            return await self._parse_csv(file_path)
        else:
            return await self._parse_excel(file_path)

    async def _parse_csv(self, file_path: Path) -> ParsedDocument:
        """Parse a CSV file."""
        metadata = self._get_base_metadata(file_path)
        metadata["language"] = "csv"

        try:
            # Detect encoding
            content_bytes = file_path.read_bytes()
            for encoding in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
                try:
                    content_str = content_bytes.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                content_str = content_bytes.decode("utf-8", errors="ignore")

            # Parse CSV
            lines = content_str.splitlines()
            reader = csv.reader(lines)
            rows = list(reader)

            if not rows:
                return ParsedDocument(
                    content="",
                    metadata=metadata,
                    sections=[],
                )

            # Format as text
            formatted_lines = []
            headers = rows[0] if rows else []

            for i, row in enumerate(rows):
                if i == 0:
                    formatted_lines.append("Headers: " + " | ".join(row))
                else:
                    # Create key-value pairs for data rows
                    pairs = []
                    for j, cell in enumerate(row):
                        if cell.strip():
                            header = headers[j] if j < len(headers) else f"Col{j}"
                            pairs.append(f"{header}: {cell}")
                    if pairs:
                        formatted_lines.append(f"Row {i}: " + ", ".join(pairs))

            content = "\n".join(formatted_lines)

            metadata["row_count"] = len(rows)
            metadata["column_count"] = len(headers)

            # Create section for the data
            sections = [
                DocumentSection(
                    title="CSV Data",
                    content=content,
                    level=0,
                )
            ]

            logger.info(f"Parsed CSV: {file_path.name} ({len(rows)} rows)")

            return ParsedDocument(
                content=content,
                metadata=metadata,
                sections=sections,
            )

        except Exception as e:
            logger.error(f"Failed to parse CSV {file_path}: {e}")
            raise RuntimeError(f"Failed to parse CSV: {e}")

    async def _parse_excel(self, file_path: Path) -> ParsedDocument:
        """Parse an Excel file."""
        metadata = self._get_base_metadata(file_path)
        metadata["language"] = "excel"

        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)

            all_content = []
            sections = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_content = []

                # Get all rows
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]

                for i, row in enumerate(rows):
                    if i == 0:
                        sheet_content.append("Headers: " + " | ".join(str(h) for h in row if h))
                    else:
                        # Create key-value pairs
                        pairs = []
                        for j, cell in enumerate(row):
                            if cell is not None and str(cell).strip():
                                header = headers[j] if j < len(headers) else f"Col{j}"
                                pairs.append(f"{header}: {cell}")
                        if pairs:
                            sheet_content.append(f"Row {i}: " + ", ".join(pairs))

                if sheet_content:
                    content = "\n".join(sheet_content)
                    all_content.append(f"Sheet: {sheet_name}\n{content}")
                    sections.append(
                        DocumentSection(
                            title=f"Sheet: {sheet_name}",
                            content=content,
                            level=1,
                        )
                    )

            wb.close()

            full_content = "\n\n".join(all_content)
            metadata["sheet_count"] = len(wb.sheetnames)

            logger.info(f"Parsed Excel: {file_path.name} ({len(wb.sheetnames)} sheets)")

            return ParsedDocument(
                content=full_content,
                metadata=metadata,
                sections=sections,
            )

        except ImportError:
            raise RuntimeError(
                "openpyxl is required for Excel parsing. "
                "Install with: pip install openpyxl"
            )
        except Exception as e:
            logger.error(f"Failed to parse Excel {file_path}: {e}")
            raise RuntimeError(f"Failed to parse Excel file: {e}")
